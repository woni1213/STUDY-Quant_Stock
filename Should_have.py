import Kiwoom_OpenAPI_Mod
import should_have_def
import sys
import time
import os
import pandas
import openpyxl
from PyQt5.QtWidgets import *
from openpyxl import Workbook
from pandas import Series, DataFrame
from datetime import datetime, timedelta
import pandas_datareader.data as web
import requests
import telegram
from bs4 import BeautifulSoup
import threading


def main_menu():
    print(should_have_def.main_menu_str)
    number = input("숫자를 입력하세요: ")  # string으로 받음

    if number == "1":
        user_menu()

    elif number == "3":
        code = input("종목 코드를 입력하세요: ")
        get_market_info(code)

    elif number == "10":
        pandas_week_candle_good_company()

    elif number == "11":
        tr_real_time_top_trade_rate()

    elif number == "12":
        test_code_name = ["aaaaaa", "bbbbb", "cccccc"]
        test_str = ""
        for name in test_code_name:
            test_str += name
            test_str += "\n"
        telgm(test_str)

    elif number == "13":
        telgm_file_send("G:\내 드라이브\학습\Python\Quant_Stock\should_have_def.py")

    # elif number == "14":
    #     order(1, "000020", 10)

    elif number == "21":
        data = my_asset("000020")
        print(data)
        main_menu()

    elif number == "22":
        my_asset_result()
        main_menu()

    elif number == "25":
        time_now = datetime.now()
        today = time_now.strftime('%Y%m%d')  # 날짜

        my_asset("0")
        trading_wb = openpyxl.load_workbook('result/Trading.xlsx')
        trading_ws = trading_wb[today]

        trading_ws_max_row = trading_ws.max_row

        sell_code = []
        sell_qty = []

        for i in range(trading_ws_max_row - 4):
            sell_code.append(trading_ws.cell(trading_ws_max_row - i, 1).value)
            sell_qty.append(int(trading_ws.cell(trading_ws_max_row - i, 3).value))

        # for i in range(len(sell_code)):
        #     order(2, sell_code[i], sell_qty[i])

        main_menu()

        # sell_code.append()
        #
        #
        # order(2, code, qty):
        #     data = KW.SendOrder(order_type, code, qty)
        #     print(data)

    elif number == "88":
        tr_get_min_candle()

    elif number == "99":
        pass

    else:
        main_menu()


def user_menu():
    print(KW.GetLoginInfo("ACCOUNT_CNT"))
    print(KW.GetLoginInfo("ACCNO"))
    print(KW.GetLoginInfo("USER_ID"))
    print(KW.GetLoginInfo("USER_NAME"))
    print(KW.GetLoginInfo("KEY_BSECGB"))
    print(KW.GetLoginInfo("FIREW_SECGB"))

    main_menu()


def kospi_dic():
    kospi = KW.GetCodeListByMarket('0')

    print(kospi)

    dic_kospi = {kospi[0]: 'kospi_len'}

    for i in kospi:
        dic_kospi[i] = KW.GetMasterCodeName(i)

    print(dic_kospi)

    return dic_kospi


def kosdaq_dic():
    kosdaq = KW.GetCodeListByMarket('10')

    dic_kosdaq = {kosdaq[0]: 'kosdaq_len'}

    for i in kosdaq:
        dic_kosdaq[i] = KW.GetMasterCodeName(i)

    print(dic_kosdaq)

    return dic_kosdaq


def get_market_info(code):
    KW.SetInputValue("종목코드", code)
    data = KW.CommRqData("opt10001", should_have_def.opt10001)

    print("")
    for i in range(len(data)):
        data_var = data[i]
        for j in range(len(data_var)):
            print(should_have_def.opt10001[j] + " : " + data_var[j])
    print("")

    main_menu()


###################

# 키움 TR요청으로 전 종목 주봉 데이터로 우상향 우량주 판별 함수
# --------------------------------------------
# 하지만 TR로 요청 시 1000번째에서 강제 종료 당함 (서버 과부하로 인해)
# 따라서 미사용

def tr_week_candle_good_company():
    stock_1 = []
    num = 1

    write_wb = Workbook()
    write_wb.create_sheet('주봉')
    write_ws = write_wb.active

    for stock_code in stock_code_name:
        print("%d / %d" % (num, len(stock_code_name)))
        print("종목 코드 : " + stock_code)
        print("종목 이름 : " + stock_name[num - 1])

        KW.SetInputValue("종목코드", stock_code)
        KW.SetInputValue("기준일자", time.strftime('%Y%m%d', time.localtime(time.time())))
        KW.SetInputValue("끝일자", "1")  # 적용 안됨 ㅅㅂ
        KW.SetInputValue("수정주가구분", "1")

        data = KW.CommRqData("opt10082", ["현재가"])

        today_cost = int(data[0][0])
        ma_60 = 0
        ma_20 = 0
        row_cost = [stock_code]

        if len(data) > 60:
            for i in range(60):
                ma_60 += int(data[i][0])
                row_cost.append(data[i][0])
        else:
            ma_60 = 0

        if len(data) > 20:
            for i in range(20):
                ma_20 += int(data[i][0])
        else:
            ma_20 = 0

        ma_60 = ma_60 / 60
        ma_20 = ma_20 / 20

        row_cost.insert(1, ma_20)
        row_cost.insert(2, ma_60)

        write_ws.append(row_cost)

        print("20일이평선 : %d" % ma_20)
        print("60일이평선 : %d" % ma_60)
        print("현재가 : %d" % today_cost)

        if today_cost > ma_20 and today_cost > ma_60:
            stock_1.append(stock_code)
            print("당첨!!!!!!!")

        print("")

        time.sleep(1)
        num += 1

    print("ToTal : %d" % len(stock_1))
    print(stock_1)

    write_wb.save("주봉.xlsx")


# Pandas로 naver 금융 싸이트의 전 종목 일봉 데이터로 우상향 우량주 판별 함수
# --------------------------------------------
# 20, 60일 이평선 위로 주가가 형성된 비율로 상위 종목 판별
# 1년 기간으로 20, 60일 이평선 이상일 경우 1 더하고
# 현재와 가까울 수록 offset으로 더 많은 수치를 더하여 %로 환산함

def pandas_week_candle_good_company():
    count = 0
    raw_data = 0
    today = datetime.now()
    winner_count = 0

    rd = {'code': [], 'name': [], '20': [], '60': []}
    winner = DataFrame(rd)

    year = today.year - 1  # 1년 전
    month = today.month
    day = today.day

    today = today.strftime('%Y-%m-%d')

    before = datetime(year, month, day)

    before = before.strftime('%Y-%m-%d')

    for stock_code in stock_code_name:  # 전 종목 반복
        offset = 0
        count += 1
        week_candle_count = 0
        month_candle_count = 0
        week_candle = []
        month_candle = []

        df = web.DataReader(stock_code, 'naver', start=before, end=today)  # naver

        day_candle = df['Low'].tolist()

        if len(day_candle) > 100:
            for i in range(len(day_candle)):  # 20일 이평선 리스트
                for j in range(20):
                    raw_data += int(day_candle[i - j])
                week_candle.append(str(raw_data // 20))
                raw_data = 0

            for i in range(20):
                week_candle[i] = '0000'  # 연산 불가 구간 0000으로 변경

            for i in range(len(day_candle)):  # 60일 이평선 리스트
                for j in range(60):
                    raw_data += int(day_candle[i - j])
                month_candle.append(str(raw_data // 60))
                raw_data = 0

            for i in range(60):  # 연산 불가 구간 0000으로 변경
                month_candle[i] = '0000'

            for i in range(len(day_candle)):  # 점수 계산
                offset += i  # 최종 연산 시 %로 변환하기 위한 offset
                if day_candle[i] > month_candle[i]:
                    month_candle_count += 1
                    month_candle_count += i

                if day_candle[i] > week_candle[i]:
                    week_candle_count += 1
                    week_candle_count += i

            month_candle_count = month_candle_count - 60
            week_candle_count = week_candle_count - 20

            # 최종 연산
            month_candle_count = (month_candle_count / (len(day_candle) - 60 + offset)) * 100
            week_candle_count = (week_candle_count / (len(day_candle) - 20 + offset)) * 100

            # 최종 판별. 80% 이상일 경우 저장
            if month_candle_count > 80 and week_candle_count > 80:
                winner_count += 1
                winner.loc[len(winner)] = [stock_code, stock_dic[stock_code], week_candle_count, month_candle_count]

        print(stock_code + "      " + str(count) + " / " + str(len(stock_code_name)) + "      " + str(winner_count))

    print(winner)

    path = './result'

    if not os.path.exists(path):
        os.makedirs(path)

        # if not os.path.exists('./data/stock_raw_data.xlsx'):
        # with pandas.ExcelWriter('./data/stock_raw_data.xlsx', mode='w', engine='openpyxl') as writer:
        # df.to_excel(writer, sheet_name=stock_code)

        # else:
        # with pandas.ExcelWriter('./data/stock_raw_data.xlsx', mode='a', engine='openpyxl') as writer:
        # df.to_excel(writer, sheet_name=stock_code)

    # count += 1
    # print(stock_code + "      " + str(count) + " / " + str(len(stock_code_name)))

    winner.to_excel('./result/' + today + '_sort_20_60.xlsx')


# Pandas로 naver 금융 싸이트의 전 종목 일봉 데이터 저장 파일의 최신 업데이트 함수
# --------------------------------------------
# 엑셀 파일을 불러와서 리스트에 넣고 난 뒤 최신 버전으로 업데이트 함
# 전 종목 일봉 데이터 저장 함수를 삭제하여 미사용

def pandas_stock_update_test():
    global old_day
    day_flag = False

    df_excel = pandas.read_excel('./data/stock_raw_data.xlsx', index_col='Date')

    today = datetime.now()

    day_count = today.day
    month_count = today.month
    year_count = today.year

    today = today.strftime('%Y-%m-%d')

    while not day_flag:
        old_day = datetime(year_count, month_count, day_count)

        day_flag = old_day in df_excel.index

        if day_flag == True:
            break

        elif day_count == 1:
            if month_count == 1:
                year_count -= 1
                month_count = 12
                day_count = 31

            elif month_count == 5 or month_count == 7 or month_count == 10 or month_count == 12:
                month_count -= 1
                day_count = 30

            elif month_count == 3:
                month_count -= 1
                day_count = 28

            else:
                month_count -= 1
                day_count = 31

        else:
            day_count -= 1

    before = old_day.strftime('%Y-%m-%d')

    if before == today:
        print("최신 자료임")

    else:
        df = web.DataReader('고쳐야러ㅑㄴ어레ㅐㅑ어냐ㅔㅐ', 'naver', start=before, end=today)

        df = df.drop(df.index[0])

        result1 = pandas.concat([df_excel, df])

        path = './data'

        if not os.path.exists(path):
            os.makedirs(path)

        result1.to_excel('./data/stock_raw_data.xlsx', sheet_name='new_name')


# 거래량 급증 항목을 받아서 처리하는 함수
# --------------------------------------------
# 3분마다 거래량 급증항목을 받아서 상위 20개 종목만 처리함
# 3분 전 거래량보다 70% 이상 증가했는지, 주가가 상승했을 경우 저장 및 출력함
# 반대로 주가가 상승 후 하락했을 경우도 저장 및 출력함
# 상승 저장한 종목은 실시간으로 주가 모니터링 함
# 만약 고점에서 1% 하락 시 출력함
# 텔레그램, 엑셀 사용

def tr_real_time_top_trade_rate():
    global min_candle_ws, min_candle_20, min_candle_60, volume_range
    telgm("실시간 상위 거래량 서치 시작")

    print("")

    plus_code_list = []  # 상승 종목 횟수 저장용
    minus_code_list = []  # 하락 종목 횟수 저장용
    sel_code_list = []  # 실시간 주가 검사용
    sel_code_history_list = []  # 상승 주가 저장용
    am_data = []
    plus_list_cnt = 0  # 조건문 내에 있어서 초기화 필요함
    upper_flag = False

    time_now = datetime.now()
    today = time_now.strftime('%Y-%m-%d')  # 날짜

    write_wb = Workbook()  # 엑셀 워크북 생성
    min_candle_wb = Workbook()  # 이평선용 워크북 생성

    while 920 > ((time_now.hour * 60) + time_now.minute) > 540:  # 오전 9시부터 오후 3시 20분까지 반복함
    # while True:
        time_now = datetime.now()  # 현재 시간

        print(time_now.strftime("%H:%M"))

        plus_list = []  # 텔레그램 출력용
        minus_list = []
        plus_list_pc = []  # PC 출력용
        minus_list_pc = []
        volume_data = []
        high_volume_data = []
        high_result = []

        high_value = tr_data_call("opt10016", should_have_def.opt10016)
        high_volume = tr_data_call("opt10024", should_have_def.opt10024)

        for item_value in high_value:
            for item_volume in high_volume:
                if item_value[0] == item_volume[0]:
                    high_result.append(item_value[0])

        for item in high_result:
            stock_name_var = stock_dic[item]
            etf = 0

            for name in should_have_def.etf:  # etf 종목 제외
                etf += stock_name_var.count(name)

            if am_data.count(item) == 0 and etf == 0:
                am_data.append(item)
                high_volume_data.append("[%s](https://finance.naver.com/item/main.nhn?code=%s)"
                                        % (stock_dic[item], item))

        if not (8 < time_now.hour < 10):  # 평상시
            volume_data = tr_data_call("opt10023", should_have_def.opt10023)

            if len(volume_data) > 21:
                volume_range = 20

            else:
                volume_range = len(volume_data)

            for i in range(volume_range):  # 거래량 상위 20개 종목만
                volume_data_var = volume_data[i]  # 리스트 내 리스트 [[상위 1 종목], [상위 2 종목], ... ]

                stock_name_var = volume_data_var[1]
                etf = 0

                for name in should_have_def.etf:  # etf 종목 제외
                    etf += stock_name_var.count(name)

                xml_data = []

                now_cost_str = volume_data_var[2]
                now_cost = int(sign_judge(now_cost_str))  # 기호 삭제
                yest_cost = now_cost - int(volume_data_var[4])  # 전일 주가

                if not now_cost == int(volume_data_var[4]):
                    cost_rate = (int(volume_data_var[4]) * 100) / yest_cost  # 주가 상승률

                else:
                    cost_rate = 0

                print("종목 이름 : %s / 상승 : %s" % (stock_name_var, volume_data_var[8]))

                if float(volume_data_var[8]) > 70 and etf == 0:  # 상승률 70% 이상, etf 종목 제외
                    min_candle_data = tr_get_min_candle(volume_data_var[0])

                    before_cost_str = min_candle_data[1][0]  # 3분전 주가 요청
                    before_cost = int(sign_judge(before_cost_str))  # 기호 삭제
                    cost_diff = now_cost - before_cost  # 분봉 대비 주가 변화 금액
                    min_rate = (before_cost // 100) * 3  # 분봉 대비 3%

                    print("종목 이름 : %s. 증가. 전 : %d / 후 : %d" % (stock_name_var, before_cost, now_cost))

                    if (before_cost + min_rate) < now_cost:  # 앞 분봉대비 3% 증가하였을 경우, 오전 9시 데이터
                        upper_flag = True
                        plus_list_cnt = plus_code_list.count(volume_data_var[0])  # 중복 횟수 카운트

                        if not plus_list_cnt == 0:
                            write_ws = write_wb[volume_data_var[1]]  # 시트 활성화
                            high_cost = int(write_ws.cell(1, 10).value)  # 상승가 데이터 읽기

                        else:
                            high_cost = 0

                        if high_cost < now_cost or plus_list_cnt == 0:
                            plus_code_list.append((volume_data_var[0]))  # 여기에 중복해서 계속 저장함

                            plus_list.append(
                                "[%s](https://finance.naver.com/item/main.nhn?code=%s) | %d - %d = %d / %.2f | %s | %d"
                                % (volume_data_var[1], volume_data_var[0], now_cost, before_cost, cost_diff, cost_rate,
                                   volume_data_var[8], plus_list_cnt))  # 텔레그램 출력용

                            plus_list_pc.append("[%s] | %d - %d = %d / %.2f | %s | %d"
                                                % (volume_data_var[1], now_cost, before_cost, cost_diff, cost_rate,
                                                   volume_data_var[8], plus_list_cnt))  # pc 모니터링 용
                            # 엑셀 저장용
                            xml_data.append(time_now.strftime("%H:%M"))
                            xml_data.append(yest_cost)
                            xml_data.append(now_cost)
                            xml_data.append(before_cost)
                            xml_data.append(cost_rate)
                            xml_data.append(cost_diff)
                            xml_data.append(volume_data_var[8])

                            if sel_code_list.count(volume_data_var[0]) == 0:  # 실시간 주가 검색용 리스트. 중복 저장이 안되게 조건문 만듬
                                sel_code_list.append(volume_data_var[0])
                                sel_code_history_list.append("%s" % volume_data_var[1])

                    else:
                        if plus_code_list.count(volume_data_var[0]) > 0:  # 거래량 상승이 1번이라도 있다면
                            if before_cost > now_cost:  # 하락추세라면
                                minus_code_list.append(volume_data_var[0])
                                minus_list_cnt = minus_code_list.count(volume_data_var[0])

                                minus_list.append(
                                    "[%s](https://finance.naver.com/item/main.nhn?code=%s) | %d - %d = %d / %.2f | %s | %d"
                                    % (volume_data_var[1], volume_data_var[0], now_cost, before_cost, cost_diff,
                                       cost_rate,
                                       volume_data_var[8],
                                       minus_list_cnt))

                                minus_list_pc.append("[%s] | %d - %d = %d / %.2f | %s | %d"
                                                     % (volume_data_var[1], now_cost, before_cost, cost_diff,
                                                        cost_rate,
                                                        volume_data_var[8],
                                                        minus_list_cnt))

                                xml_data.append(time_now.strftime("%H:%M"))
                                xml_data.append(yest_cost)
                                xml_data.append(now_cost)
                                xml_data.append(before_cost)
                                xml_data.append(cost_rate)
                                xml_data.append(cost_diff)
                                xml_data.append(volume_data_var[8])

                                if sel_code_list.count(volume_data_var[0]) > 0:
                                    sel_code_list.remove(volume_data_var[0])

                                write_ws = write_wb[volume_data_var[1]]  # 활성화

                                # sell_qty = my_asset(volume_data_var[0])

                                # if not sell_qty == 0:
                                #     order(2, volume_data_var[0], sell_qty)

                    if not len(xml_data) == 0:  # 만약 엑셀 저장용 데이터가 존재한다면
                        if plus_list_cnt == 0:  # 처음 등록하는 거라면
                            write_wb.create_sheet(volume_data_var[1])  # 종목 이름으로 시트 생성
                            write_ws = write_wb[volume_data_var[1]]  # 활성화
                            write_ws.append(should_have_def.xml_data_name)  # 항목 이름 추가

                            # order(1, volume_data_var[0], (1000000 // now_cost))

                            min_candle_wb.create_sheet(volume_data_var[0])  # 종목 코드로 시트 생성

                        write_ws = write_wb[volume_data_var[1]]  # 시트 활성화

                        if upper_flag:
                            min_candle_wb.remove(min_candle_wb[volume_data_var[0]])  # 시트 삭제 후 재생성
                            min_candle_wb.create_sheet(volume_data_var[0])

                            # min_candle_ws = min_candle_wb[volume_data_var[0]]
                            #
                            # if len(min_candle_data) > 119:
                            #     for j in range(120):  # 최대 3분봉 120개 저장
                            #         min_candle_list = []
                            #         min_candle_str = min_candle_data[119 - j][0]  # 역순 저장
                            #         min_candle_str = sign_judge(min_candle_str)
                            #         min_candle_list.append(min_candle_str)
                            #         min_candle_ws.append(min_candle_list)

                        write_ws.cell(1, 10, now_cost)  # 상승가 저장용
                        write_ws.cell(1, 11, 0)  # 가격 하락 Flag
                        # min_candle_ws.cell(1, 2, 0)  # 20 이평 Flag
                        # min_candle_ws.cell(1, 3, 0)  # 60 이평 Flag

                        upper_flag = False
                        write_ws.append(xml_data)

            print("조건식 완료")
            print("")

            telgm_plus_list = "---- 거래량 대비 상승 종목 ----\n"
            telgm_minus_list = "---- 거래량 대비 하락 종목 ----\n"

            for telgm_name in plus_list:
                telgm_plus_list += telgm_name
                telgm_plus_list += "\n"

            for telgm_name in minus_list:
                telgm_minus_list += telgm_name
                telgm_minus_list += "\n"

            print('\033[31m' + "---- 거래량 대비 상승 종목 ----" + '\033[0m')

            if not len(plus_list) == 0:
                telgm(telgm_plus_list)

            for i in range(len(plus_list)):
                print(plus_list_pc[i])

            print("")

            print('\033[34m' + "---- 거래량 대비 하락 종목 ----" + '\033[0m')

            if not len(minus_list) == 0:
                telgm(telgm_minus_list)

            for i in range(len(minus_list)):
                print(minus_list_pc[i])

            print("")

            t_end = time.time() + 60 * 3  # 3분 타이머 선언
            # t_end = time.time() + 20  # 테스트 타이머 선언
            print("3분 타이머 시작")
            print("")
            print("-- 상세 검색 종목 --")

            for code_to_name in sel_code_list:
                print("%s" % stock_dic[code_to_name])

            while time.time() < t_end:  # 3분 타이머 실행
                for code in sel_code_list:
                    write_ws = write_wb[stock_dic[code]]
                    min_candle_ws = min_candle_wb[code]

                    high_cost = int(write_ws.cell(1, 10).value)  # 상승가 데이터 읽기
                    deg_flag = int(write_ws.cell(1, 11).value)  # 하락 신호 Flag 읽기
                    # min_candle_20_flag = int(min_candle_ws.cell(1, 2).value)  # 20 이평 Flag 읽기
                    # min_candle_60_flag = int(min_candle_ws.cell(1, 3).value)  # 20 이평 Flag 읽기
                    real_cost = int(get_price(code))  # 현재 주가 읽기 (Pandas로 네이버 크롤링)

                    # if real_cost > high_cost:           # 현재주가가 고점이라면 엑셀에 데이터 저장
                    #     write_ws.cell(1, 10, real_cost)

                    low_cost = high_cost - ((high_cost // 100) * 1)  # 상승가 에서 1% 하락폭

                    # min_candle_raw_data = 0

                    # min_candle_ws_max_row = min_candle_ws.max_row
                    #
                    # for i in range(0, 60):
                    #     min_candle_raw_data += int(
                    #         min_candle_ws.cell((min_candle_ws_max_row - i), 1).value)  # 3분봉 데이터 읽기
                    #
                    #     if i == 19:  # 3분봉 20 이평
                    #         min_candle_20 = min_candle_raw_data // 20
                    #
                    #     if i == 59:  # 3분봉 60 이평
                    #         min_candle_60 = min_candle_raw_data // 60

                    if real_cost < low_cost and deg_flag == 0:  # 상승가에서 1% 이상 하락한다면 and 하락 신호를 안보냈으면
                        xml_data = []  # 초기화

                        time_now = datetime.now()

                        xml_data.append(time_now.strftime("%H:%M"))
                        xml_data.append("하락")
                        xml_data.append(real_cost)
                        xml_data.append(high_cost)
                        xml_data.append("")
                        xml_data.append(high_cost - real_cost)
                        xml_data.append("")

                        write_ws.append(xml_data)
                        write_ws.cell(1, 11, 1)  # 하락 신호 Flag

                        print(" %s - 하락 전환" % stock_dic[code])
                        print("현재가 : %d | 상승가 : %d | 가격 차이 : %d" % (real_cost, high_cost, (high_cost - real_cost)))
                        telgm(
                            "[%s](https://finance.naver.com/item/main.nhn?code=%s) - 하락 전환 | 고점 : %d 현재 : %d 주가차이 : %d"
                            % (stock_dic[code], code, high_cost, real_cost, (high_cost - real_cost)))

                        if sel_code_list.count(code) > 0:
                            sel_code_list.remove(code)

                        write_ws = write_wb[stock_dic[code]]  # 활성화

                        # sell_qty = my_asset(code)

                        # if not sell_qty == 0:
                        #     order(2, code, sell_qty)

                    # if (real_cost < min_candle_20 and min_candle_20_flag == 0) and deg_flag == 0:
                    #     xml_data = []  # 초기화
                    #
                    #     time_now = datetime.now()
                    #
                    #     xml_data.append(time_now.strftime("%H:%M"))
                    #     xml_data.append("20이평 하락")
                    #     xml_data.append(real_cost)
                    #     xml_data.append(min_candle_20)
                    #     xml_data.append("")
                    #     xml_data.append("")
                    #     xml_data.append("")
                    #
                    #     write_ws.append(xml_data)
                    #     min_candle_ws.cell(1, 2, 1)  # 20 이평 신호 Flag
                    #
                    #     print(" %s - 20이평 하방" % (stock_dic[code]))
                    #     print("현재가 : %d | 20이평 : %d" % (real_cost, min_candle_20))
                    #
                    #     telgm("[%s](https://finance.naver.com/item/main.nhn?code=%s) - 20이평 하방 | 20이평 : %d 현재 : %d"
                    #           % (stock_dic[code], code, min_candle_20, real_cost))
                    #
                    # if (real_cost < min_candle_60 and min_candle_60_flag == 0) and deg_flag == 0:
                    #     xml_data = []  # 초기화
                    #
                    #     time_now = datetime.now()
                    #
                    #     xml_data.append(time_now.strftime("%H:%M"))
                    #     xml_data.append("60이평 하락")
                    #     xml_data.append(real_cost)
                    #     xml_data.append(min_candle_60)
                    #     xml_data.append("")
                    #     xml_data.append("")
                    #     xml_data.append("")
                    #
                    #     write_ws.append(xml_data)
                    #     min_candle_ws.cell(1, 3, 1)  # 20 이평 신호 Flag
                    #
                    #     print(" %s - 60이평 하방" % (stock_dic[code]))
                    #     print("현재가 : %d | 60이평 : %d" % (real_cost, min_candle_60))
                    #
                    #     telgm("[%s](https://finance.naver.com/item/main.nhn?code=%s) - 60이평 하방 | 60이평 : %d 현재 : %d"
                    #           % (stock_dic[code], code, min_candle_60, real_cost))

            # for code in sel_code_list:  # 현재가 3분봉 엑셀에 추가
            #     min_candle_list = []
            #     min_candle_ws = min_candle_wb[code]
            #
            #     real_cost = int(get_price(code))  # 현재 주가 읽기 (Pandas로 네이버 크롤링)
            #
            #     min_candle_list.append(real_cost)
            #     min_candle_ws.append(min_candle_list)  # 엑셀에 현재 주가 추가

        telgm_am_list = "---- 전일 거래량 100% 및 52주 최고가 돌파 종목 ----\n"

        for telgm_name in high_volume_data:
            telgm_am_list += telgm_name
            telgm_am_list += "\n"

        if not len(high_volume_data) == 0:
            telgm(telgm_am_list)

        if 8 < time_now.hour < 10:  # 9시 부터 10시까지
            time.sleep(60)

        print("타이머 종료")
        print("")

    path = './result'

    if not os.path.exists(path):
        os.makedirs(path)

    path_1 = './result/' + today + '_Trading_Volume.xlsx'
    path_2 = './result/' + today + '_min_candle_data.xlsx'

    write_wb.save(path_1)
    min_candle_wb.save(path_2)

    telgm("장마감")

    main_menu()


# def order(order_type, code, qty):
#     data = KW.SendOrder(order_type, code, qty)


def tr_data_call(tr_command, tr_name):
    while True:
        KW.SetInputValue(tr_name[0], tr_name[1])

        data = KW.CommRqData(tr_command, tr_name[2])  # 거래량 증가 종목 요청

        if not data == 0:
            break

    return data


def sign_judge(data):
    if data.count('-') or data.count('+'):
        data = (data[1:])  # 기호 삭제

    return data


def tr_get_min_candle(code):
    min_candle_data_raw = should_have_def.opt10080
    code_data_raw = min_candle_data_raw.pop(1)
    code_data_raw.insert(0, code)
    min_candle_data_raw.insert(1, code_data_raw)

    data = tr_data_call("opt10080", min_candle_data_raw)

    return data


def get_price(code):
    url = "https://finance.naver.com/item/main.nhn?code=" + code
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")
    no_today = bs_obj.find("p", {"class": "no_today"})
    blind_now = no_today.find("span", {"class": "blind"})

    return blind_now.text.replace(',', '')


def telgm(telgm_text):
    bot = telegram.Bot(token=should_have_def.bot_token)
    bot.sendMessage(chat_id=should_have_def.group_id, text=telgm_text, parse_mode="Markdown",
                    disable_web_page_preview=True)


def telgm_file_send(file_path):
    bot = telegram.Bot(token=should_have_def.bot_token)
    bot.sendDocument(chat_id=should_have_def.solo_id, document=open(file_path, 'rb'))


def tr_test():
    data = KW.CommRqData("", should_have_def.opt10023)
    print(data)


def num_to_cost_type(num):
    if num.count(".") == 0:
        data = int(num)
        data = "{:,}".format(data)

    else:
        data = float(num)

    return data


def my_asset(code):
    time_now = datetime.now()
    today = time_now.strftime('%Y%m%d')  # 날짜

    trading_wb = openpyxl.load_workbook('result/Trading.xlsx')
    trading_ws_title = trading_wb.sheetnames

    if trading_ws_title.count("trading_" + today) == 0:
        trading_wb.create_sheet("trading_" + today)

    else:
        trading_wb.remove(trading_wb["trading_" + today])  # 시트 삭제 후 재생성
        trading_wb.create_sheet("trading_" + today)

    trading_ws = trading_wb["trading_" + today]

    data = tr_data_call("OPW00004", should_have_def.OPW00004)

    asset = num_to_cost_type(data[0][0])  # 예탁자산평가액
    margin = num_to_cost_type(data[0][1])  # 당일손익률

    trading_ws.cell(1, 1, "예탁 자산 평가액")
    trading_ws.cell(2, 1, "당일 손익")

    trading_ws.cell(1, 2, asset)
    trading_ws.cell(2, 2, margin)

    trading_ws.append([""])

    menu = should_have_def.OPW00004[2]
    menu = menu[2:12]

    trading_ws.append(menu)

    if not data[0][11] == "0000":
        for i in range(len(data)):
            code_list = data[i][2]
            code_list = code_list[1:]  # 종목 코드
            name_list = data[i][3]  # 종목 이름
            qty = str(int(data[i][4]))  # 보유 수량
            avr_price = str(float(data[i][5]))  # 평균 단가
            real_price = str(int(data[i][6]))  # 현재가
            total_price = str(num_to_cost_type(data[0][7]))  # 평가 금액
            diff_price = str(num_to_cost_type(data[0][8]))  # 손익 금액
            diff_rate = str(float(data[i][9]))  # 손익률
            buy_price = str(num_to_cost_type(data[0][10]))  # 매입 금액

            stock_data = [code_list, name_list, qty, avr_price, real_price, total_price, diff_price, diff_rate,
                          buy_price]
            trading_ws.append(stock_data)

    trading_wb.save('result/Trading.xlsx')

    if not code == "0":
        flag = False

        for row in trading_ws.values:
            data_var = []
            for cell in row:
                data_var.append(cell)

            if data_var.count(code) == 1:
                flag = True
                break

        if flag:
            return int(data_var[2])

        else:
            return 0

    # 평가금액이상함


def my_asset_result():
    time_now = datetime.now()
    today = time_now.strftime('%Y%m%d')  # 날짜

    trading_wb = openpyxl.load_workbook('result/Trading.xlsx')
    trading_ws_title = trading_wb.sheetnames

    if trading_ws_title.count("result_" + today) == 0:
        trading_wb.create_sheet("result_" + today)

    else:
        trading_wb.remove(trading_wb["result_" + today])  # 시트 삭제 후 재생성
        trading_wb.create_sheet("result_" + today)

    trading_ws = trading_wb["result_" + today]

    data = tr_data_call("OPW00007", should_have_def.OPW00007)

    trading_ws.append(should_have_def.OPW00007[2])

    for i in range(len(data)):
        data_list = data[i]
        data_list[3] = int(data_list[3])
        data_list[4] = int(data_list[4])
        data_list.append(int(data_list[3]) * int(data_list[4]))

        trading_ws.append(data_list)

    trading_wb.save('result/Trading.xlsx')


if not QApplication.instance():  # Open API+ 인스턴스가 열리지 않았다면
    app = QApplication(sys.argv)  # 객체를 실행하기 위한 절대경로(sys.argv) 지정

KW = Kiwoom_OpenAPI_Mod.Kiwoom_Quant()

KW.Login(block=True)

print("로그인 완료")

kospi = kospi_dic()
kosdaq = kosdaq_dic()

stock_dic = {}

stock_dic.update(kospi)
stock_dic.update(kosdaq)

# stock_code_name = list(kospi.keys()) + list(kosdaq.keys())
# stock_name = list(kospi.values()) + list(kosdaq.values())

stock_code_name = list(stock_dic.keys())
stock_name = list(stock_dic.values())

print(len(stock_code_name))

main_menu()

KW.Logout()

print("형산강 가자")

# pip install python-telegram-bot

# 6.10
# 예외처리
# 전날 구매했다가 안팔린 종목 일딴 당일 시작하면 다 매도
# 9시10시 조건식에 ETF 제외
# 너무 고점에서 잡는다 이거. 1분봉으로 변경해서 테스트