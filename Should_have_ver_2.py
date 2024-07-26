import Kiwoom_OpenAPI_Mod
import should_have_def_2
import sys
import time
import os
import pandas
import openpyxl
from PyQt5.QtWidgets import *
from openpyxl import Workbook
from pandas import Series, DataFrame
from datetime import datetime, timedelta
import pandas_datareader.data as web
import requests
import telegram
import threading
from bs4 import BeautifulSoup
from pathlib import Path
from PyQt5.QAxContainer import *
import pythoncom  # COM 관련

def main_menu():
    print(should_have_def_2.main_menu_str)
    number = input("숫자를 입력하세요: ")  # string으로 받음

    if number == "1":
        thread_str = threading.Thread(target=thread_breaking_high_search)
        thread_start(thread_str, number)

        main_menu()

    elif number == "2":
        thread_str = threading.Thread(target=thread_breaking_high_search)
        thread_start(thread_str, number)

        main_menu()

    elif number == "99":
        pass

    else:
        main_menu()


def thread_start(thread_func, flag):
    thread_flag_key = thread_flag.get(flag)

    if (thread_flag_key is None) or thread_flag_key == 0:
        thread_func.daemon = True
        thread_flag[flag] = 1
        thread_func.start()

    elif thread_flag_key == 1:
        thread_flag[flag] = 0


###################

# 거래량 급증 항목을 받아서 처리하는 함수
# --------------------------------------------
# 3분마다 거래량 급증항목을 받아서 상위 20개 종목만 처리함
# 3분 전 거래량보다 70% 이상 증가했는지, 주가가 상승했을 경우 저장 및 출력함
# 반대로 주가가 상승 후 하락했을 경우도 저장 및 출력함
# 상승 저장한 종목은 실시간으로 주가 모니터링 함
# 만약 고점에서 1% 하락 시 출력함
# 텔레그램, 엑셀 사용

def thread_breaking_high_search():
    global shared_cmd
    global tr_data
    telegram_high_stock = []

    while thread_flag["1"] == 1:  # 쓰레드 런, 스탑
        # while True:
        stock_time = time_now(0)

        # if 920 > ((stock_time.hour * 60) + stock_time.minute) > 540:  # 오전 9시부터 오후 3시 20분까지 반복함
        if True:
            print(time_now(1))

            high_stock = []
            high_stock_ws = high_stock_wb[time_now(2)]

            high_stock_ws_max_row = high_stock_ws.max_row - 1

            if not high_stock_ws_max_row == 0:
                for i in range(high_stock_ws_max_row):
                    high_stock.append(high_stock_ws.cell(i + 1, 1).value)

            xml_write_renew_ws(high_stock_wb, [time_now(2)])
            high_stock_ws = high_stock_wb[time_now(2)]

            lock.acquire()

            shared_cmd = "opt10016"
            while not shared_cmd == "":
                pass
            high_value = tr_data

            shared_cmd = "opt10024"
            while not shared_cmd == "":
                pass
            high_volume = tr_data

            lock.release()

            for item_value in high_value:
                for item_volume in high_volume:
                    if item_value[0] == item_volume[0]:
                        high_stock.append(item_value[0])

            high_stock = set_data(high_stock)

            for item in high_stock:
                high_stock_ws.append([item])

            high_stock_var = set_data_cal(1, high_stock, telegram_high_stock)
            telegram_high_stock = telegram_high_stock + high_stock_var

            high_stock_var = telegram_link(high_stock_var)

            if not len(high_stock_var) == 0:
                telegram_msg("---- 전일 거래량 100% 및 52주 최고가 돌파 종목 ----", high_stock_var)

            time.sleep(5)

        else:
            thread_flag["1"] = 0

    high_stock_wb.save('result/high_stock.xlsx')


def telegram_msg(str_1, str_2):
    telegram_msg_send = str_1
    telegram_msg_send += "\n"

    for name in str_2:
        telegram_msg_send += name
        telegram_msg_send += "\n"

    telegram_text(telegram_msg_send)


def telegram_link(data):
    var = []
    for item in data:
        var.append("[%s](https://finance.naver.com/item/main.nhn?code=%s)"
                   % (stock_dic[item], item))
    return var


def set_data_cal(cmd, data_1, data_2):
    s1 = set(data_1)
    s2 = set(data_2)

    if cmd == 1:  # 차집합
        return list(s1 - s2)


def set_data(data):
    data_set = set(data)
    data = list(data_set)

    return data


def time_now(time_type):
    data = datetime.now()

    if time_type == 0:  # timestamp
        return data

    elif time_type == 1:  # 시:분
        return data.strftime("%H:%M")

    elif time_type == 2:
        return data.strftime('%Y%m%d')

    else:
        return 0


def order(order_type, code, qty):
    data = KW.SendOrder(order_type, code, qty)


def tr_data_call(tr_command, tr_name):
    while True:
        KW.SetInputValue(tr_name[0], tr_name[1])

        data = KW.CommRqData(tr_command, tr_name[2])  # 거래량 증가 종목 요청

        if not data == 0:
            break

    return data


def sign_judge(data):
    if data.count('-') or data.count('+'):
        data = (data[1:])  # 기호 삭제

    return data


def tr_get_min_candle(code):
    min_candle_data_raw = should_have_def.opt10080
    code_data_raw = min_candle_data_raw.pop(1)
    code_data_raw.insert(0, code)
    min_candle_data_raw.insert(1, code_data_raw)

    data = tr_data_call("opt10080", min_candle_data_raw)

    return data


def get_price(code):
    url = "https://finance.naver.com/item/main.nhn?code=" + code
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")
    no_today = bs_obj.find("p", {"class": "no_today"})
    blind_now = no_today.find("span", {"class": "blind"})

    return blind_now.text.replace(',', '')


def telegram_text(msg):
    bot = telegram.Bot(token=should_have_def_2.bot_token)
    bot.sendMessage(chat_id=should_have_def_2.solo_id, text=msg, parse_mode="Markdown",
                    disable_web_page_preview=True)


def telgm_file_send(file_path):
    bot = telegram.Bot(token=should_have_def.bot_token)
    bot.sendDocument(chat_id=should_have_def.solo_id, document=open(file_path, 'rb'))


def tr_test():
    data = KW.CommRqData("", should_have_def.opt10023)
    print(data)


def num_to_cost_type(num):
    if num.count(".") == 0:
        data = int(num)
        data = "{:,}".format(data)

    else:
        data = float(num)

    return data


def my_asset(code):
    time_now = datetime.now()
    today = time_now.strftime('%Y%m%d')  # 날짜

    trading_wb = openpyxl.load_workbook('result/Trading.xlsx')
    trading_ws_title = trading_wb.sheetnames

    if trading_ws_title.count("trading_" + today) == 0:
        trading_wb.create_sheet("trading_" + today)

    else:
        trading_wb.remove(trading_wb["trading_" + today])  # 시트 삭제 후 재생성
        trading_wb.create_sheet("trading_" + today)

    trading_ws = trading_wb["trading_" + today]

    data = tr_data_call("OPW00004", should_have_def.OPW00004)

    asset = num_to_cost_type(data[0][0])  # 예탁자산평가액
    margin = num_to_cost_type(data[0][1])  # 당일손익률

    trading_ws.cell(1, 1, "예탁 자산 평가액")
    trading_ws.cell(2, 1, "당일 손익")

    trading_ws.cell(1, 2, asset)
    trading_ws.cell(2, 2, margin)

    trading_ws.append([""])

    menu = should_have_def.OPW00004[2]
    menu = menu[2:12]

    trading_ws.append(menu)

    if not data[0][11] == "0000":
        for i in range(len(data)):
            code_list = data[i][2]
            code_list = code_list[1:]  # 종목 코드
            name_list = data[i][3]  # 종목 이름
            qty = str(int(data[i][4]))  # 보유 수량
            avr_price = str(float(data[i][5]))  # 평균 단가
            real_price = str(int(data[i][6]))  # 현재가
            total_price = str(num_to_cost_type(data[0][7]))  # 평가 금액
            diff_price = str(num_to_cost_type(data[0][8]))  # 손익 금액
            diff_rate = str(float(data[i][9]))  # 손익률
            buy_price = str(num_to_cost_type(data[0][10]))  # 매입 금액

            stock_data = [code_list, name_list, qty, avr_price, real_price, total_price, diff_price, diff_rate,
                          buy_price]
            trading_ws.append(stock_data)

    trading_wb.save('result/Trading.xlsx')

    if not code == "0":
        flag = False

        for row in trading_ws.values:
            data_var = []
            for cell in row:
                data_var.append(cell)

            if data_var.count(code) == 1:
                flag = True
                break

        if flag:
            return int(data_var[2])

        else:
            return 0

    # 평가금액이상함


def my_asset_result():
    time_now = datetime.now()
    today = time_now.strftime('%Y%m%d')  # 날짜

    trading_wb = openpyxl.load_workbook('result/Trading.xlsx')
    trading_ws_title = trading_wb.sheetnames

    if trading_ws_title.count("result_" + today) == 0:
        trading_wb.create_sheet("result_" + today)

    else:
        trading_wb.remove(trading_wb["result_" + today])  # 시트 삭제 후 재생성
        trading_wb.create_sheet("result_" + today)

    trading_ws = trading_wb["result_" + today]

    data = tr_data_call("OPW00007", should_have_def.OPW00007)

    trading_ws.append(should_have_def.OPW00007[2])

    for i in range(len(data)):
        data_list = data[i]
        data_list[3] = int(data_list[3])
        data_list[4] = int(data_list[4])
        data_list.append(int(data_list[3]) * int(data_list[4]))

        trading_ws.append(data_list)

    trading_wb.save('result/Trading.xlsx')


def kospi_dic():
    kospi = KW.GetCodeListByMarket('0')

    dic_kospi = {kospi[0]: 'kospi_len'}

    for i in kospi:
        dic_kospi[i] = KW.GetMasterCodeName(i)

    print(dic_kospi)

    return dic_kospi


def kosdaq_dic():
    kosdaq = KW.GetCodeListByMarket('10')

    dic_kosdaq = {kosdaq[0]: 'kosdaq_len'}

    for i in kosdaq:
        dic_kosdaq[i] = KW.GetMasterCodeName(i)

    print(dic_kosdaq)

    return dic_kosdaq


def xml_write_wb(wb_path):
    if not Path(wb_path).is_file():
        write_wb = Workbook()
        write_wb.save('result/high_stock.xlsx')

    write_wb = openpyxl.load_workbook('result/high_stock.xlsx')

    return write_wb


def xml_write_renew_ws(write_wb, sheet_name):
    for item in sheet_name:
        if write_wb.sheetnames.count(item) == 0:
            write_wb.create_sheet(item)

        else:
            write_wb.remove(write_wb[item])  # 시트 삭제 후 재생성
            write_wb.create_sheet(item)

    if write_wb.sheetnames.count("Sheet") == 1:
        write_wb.remove(write_wb["Sheet"])

    # return write_wb


if __name__ == "__main__":
    if not QApplication.instance():  # Open API+ 인스턴스가 열리지 않았다면
        app = QApplication(sys.argv)  # 객체를 실행하기 위한 절대경로(sys.argv) 지정

    KW = Kiwoom_OpenAPI_Mod.Kiwoom_Quant()
    shared_cmd = ""
    tr_data = []

    KW.Login(block=True)

    print("로그인 완료")

    kospi = kospi_dic()
    kosdaq = kosdaq_dic()

    stock_dic = {}

    stock_dic.update(kospi)
    stock_dic.update(kosdaq)

    stock_code_name = list(stock_dic.keys())
    stock_name = list(stock_dic.values())

    print(len(stock_code_name))

    lock = threading.Lock()  # threading에서 Lock 함수 가져오기
    thread_flag = {}

    high_stock_wb = xml_write_wb('result/high_stock.xlsx')
    xml_write_renew_ws(high_stock_wb, [time_now(2)])

    main_thread = threading.Thread(target=main_menu)
    main_thread.daemon = True

    main_thread.start()

    while True:
        if not shared_cmd == "":
            if shared_cmd == "opt10016":
                tr_data = tr_data_call("opt10016", should_have_def_2.opt10016)  # 신고가

            elif shared_cmd == "opt10024":
                tr_data = tr_data_call("opt10024", should_have_def_2.opt10024)  # 5일 최고 거래량 갱신

            shared_cmd = ""

    KW.Logout()

    print("형산강 가자")

# pip install python-telegram-bot

# 6.10
# 예외처리
# 전날 구매했다가 안팔린 종목 일딴 당일 시작하면 다 매도
